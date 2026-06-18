from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── colour palette ────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E74B5"
LIGHT_BLUE = "D6E4F0"
ACCENT     = "C5E0F5"
WHITE      = "FFFFFF"
YELLOW     = "FFF2CC"
GREEN_HEAD = "E2EFDA"
ORANGE     = "FCE4D6"

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def bold_font(size=11, color="000000", name="Arial"):
    return Font(bold=True, size=size, color=color, name=name)

def norm_font(size=10, color="000000", name="Arial"):
    return Font(size=size, color=color, name=name)

def apply_header_row(ws, row_num, values, col_widths=None,
                     bg=DARK_BLUE, fg=WHITE):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.font      = bold_font(11, fg)
        c.fill      = hdr_fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = thin_border()

def apply_data_row(ws, row_num, values, alt=False):
    bg = LIGHT_BLUE if alt else WHITE
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.font      = norm_font()
        c.fill      = hdr_fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border    = thin_border()


# ═══════════════════════════════════════════════════════════════════
# SHEET 1 – Cover / Project Overview
# ═══════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Cover"
ws0.sheet_view.showGridLines = False

ws0.merge_cells("A1:H3")
c = ws0["A1"]
c.value     = "TATA SUPPLY CHAIN CO. LTD. — DATABASE MANAGEMENT SYSTEM"
c.font      = Font(bold=True, size=18, color=WHITE, name="Arial")
c.fill      = hdr_fill(DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")

ws0.row_dimensions[1].height = 30
ws0.row_dimensions[2].height = 20
ws0.row_dimensions[3].height = 20

info = [
    ("College",     "USM's Shriram Mantri Vidyanidhi Info Tech Academy"),
    ("Subject",     "Database Management System (DBMS)"),
    ("Project",     "Tata Supply Chain Co. Ltd. – Industrial DBMS"),
    ("Database",    "tata_supply_chain  |  Engine: MySQL 8.x"),
    ("Tables",      "13 Normalized Tables (3NF)"),
    ("Entities",    "Part Category, Part, Vendor, Vendor_Part, Transporter, Purchase Order, Challan, GRR, GRR_Detail, Quality Inspection, MIR, MIR_Detail"),
]
ws0.merge_cells("A4:H4")
ws0["A4"].value = ""

row = 5
for label, val in info:
    ws0.merge_cells(f"A{row}:B{row}")
    ws0.merge_cells(f"C{row}:H{row}")
    c1 = ws0[f"A{row}"]
    c1.value     = label
    c1.font      = bold_font(11, WHITE)
    c1.fill      = hdr_fill(MID_BLUE)
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.border    = thin_border()
    c2 = ws0[f"C{row}"]
    c2.value     = val
    c2.font      = norm_font(11)
    c2.fill      = hdr_fill(LIGHT_BLUE)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border    = thin_border()
    ws0.row_dimensions[row].height = 25
    row += 1

for col in ["A","B","C","D","E","F","G","H"]:
    ws0.column_dimensions[col].width = 16
ws0.column_dimensions["C"].width = 60


# ═══════════════════════════════════════════════════════════════════
# SHEET 2 – ER Entities Summary
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Entity Summary")
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:F1")
t = ws1["A1"]
t.value     = "ENTITY IDENTIFICATION & RELATIONSHIPS"
t.font      = bold_font(14, WHITE)
t.fill      = hdr_fill(DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

apply_header_row(ws1, 2,
    ["#","Entity / Table","Alias","Type","Primary Key","Relationships"],
    bg=MID_BLUE)
ws1.row_dimensions[2].height = 22

entities = [
    (1,  "Part_Category",        "PC",  "Master",      "category_id (AUTO_INCREMENT)", "One-to-Many → Part"),
    (2,  "Part",                 "P",   "Master",      "part_no (VARCHAR)",            "Many-to-One ← Part_Category; One-to-Many → Vendor_Part, GRR_Detail, MIR_Detail"),
    (3,  "Vendor",               "V",   "Master",      "vendor_code (VARCHAR)",        "One-to-Many → Vendor_Part, Purchase_Order, Challan"),
    (4,  "Vendor_Part",          "VP",  "Junction M:N","(vendor_code, part_no)",       "Many-to-One ← Vendor; Many-to-One ← Part"),
    (5,  "Transporter",          "T",   "Master",      "transporter_id (VARCHAR)",     "One-to-Many → Challan"),
    (6,  "Purchase_Order",       "PO",  "Transaction", "po_number (VARCHAR)",          "Many-to-One ← Vendor; One-to-Many → Purchase_Order_Detail, Challan"),
    (7,  "Purchase_Order_Detail","POD", "Transaction", "po_detail_id (AUTO_INCREMENT)","Many-to-One ← Purchase_Order; Many-to-One ← Part"),
    (8,  "Challan",              "CH",  "Transaction", "challan_no (VARCHAR)",         "Many-to-One ← Vendor, Transporter, Purchase_Order; One-to-Many → GRR"),
    (9,  "GRR",                  "G",   "Transaction", "grr_no (AUTO_INCREMENT)",      "Many-to-One ← Challan; One-to-Many → GRR_Detail"),
    (10, "GRR_Detail",           "GD",  "Transaction", "grr_detail_id (AUTO_INCREMENT)","Many-to-One ← GRR, Part; One-to-Many → Quality_Inspection"),
    (11, "Quality_Inspection",   "QI",  "Transaction", "inspection_id (AUTO_INCREMENT)","Many-to-One ← GRR_Detail"),
    (12, "MIR",                  "M",   "Transaction", "mir_no (VARCHAR)",             "One-to-Many → MIR_Detail"),
    (13, "MIR_Detail",           "MD",  "Transaction", "mir_detail_id (AUTO_INCREMENT)","Many-to-One ← MIR, Part"),
]

for i, row_data in enumerate(entities):
    apply_data_row(ws1, i+3, row_data, alt=(i%2==0))

widths = [4, 26, 8, 14, 30, 50]
for col, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
for r in range(3, len(entities)+4):
    ws1.row_dimensions[r].height = 22


# ═══════════════════════════════════════════════════════════════════
# helper – build a full table-definition sheet
# ═══════════════════════════════════════════════════════════════════
def make_table_sheet(wb, sheet_name, table_title, columns, pks, fks):
    """
    columns : list of (col_name, data_type, nullable, default, description)
    pks     : list of pk column names
    fks     : list of (col_name, ref_table, ref_col)
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    h = ws["A1"]
    h.value     = f"TABLE: {table_title}"
    h.font      = bold_font(13, WHITE)
    h.fill      = hdr_fill(DARK_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    apply_header_row(ws, 2,
        ["Column Name","Data Type","Nullable","Default","Key","Description","Constraints"],
        bg=MID_BLUE)
    ws.row_dimensions[2].height = 22

    pk_set  = set(pks)
    fk_dict = {f[0]: f for f in fks}

    for i, col in enumerate(columns):
        col_name, dtype, nullable, default, desc = col
        key_info = ""
        if col_name in pk_set:
            key_info = "PK"
        if col_name in fk_dict:
            key_info = ("PK, FK" if key_info == "PK" else "FK")

        constraint = ""
        if col_name in fk_dict:
            _, ref_t, ref_c = fk_dict[col_name]
            constraint = f"FK → {ref_t}({ref_c})"

        row_vals = (col_name, dtype, nullable, default, key_info, desc, constraint)
        apply_data_row(ws, i+3, row_vals, alt=(i%2==0))

        # highlight PK gold
        if col_name in pk_set:
            for c in range(1, 8):
                ws.cell(i+3, c).fill = hdr_fill(YELLOW)
        # highlight FK orange
        elif col_name in fk_dict:
            for c in range(1, 8):
                ws.cell(i+3, c).fill = hdr_fill(ORANGE)

        ws.row_dimensions[i+3].height = 20

    # FK legend
    legend_row = len(columns) + 4
    ws.merge_cells(f"A{legend_row}:G{legend_row}")
    ws[f"A{legend_row}"].value = "KEY: Yellow = Primary Key    Orange = Foreign Key"
    ws[f"A{legend_row}"].font  = bold_font(9, "444444")
    ws[f"A{legend_row}"].fill  = hdr_fill("F2F2F2")
    ws[f"A{legend_row}"].alignment = Alignment(horizontal="left")

    widths_t = [22, 26, 10, 16, 10, 38, 34]
    for col_i, w in enumerate(widths_t, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w
    return ws


# ═══════════════════════════════════════════════════════════════════
# Build individual table sheets
# ═══════════════════════════════════════════════════════════════════

make_table_sheet(wb, "Part_Category", "Part_Category",
    columns=[
        ("category_id",   "INT AUTO_INCREMENT","NOT NULL","—",     "Unique category identifier"),
        ("category_name", "VARCHAR(100)",       "NOT NULL","—",     "Name of the part category"),
        ("category_type", "ENUM(RAW/FINISHED)", "NOT NULL","—",     "Distinguishes raw vs finished"),
        ("description",   "TEXT",               "NULL",   "NULL",  "Optional detailed description"),
    ],
    pks=["category_id"], fks=[])

make_table_sheet(wb, "Part", "Part",
    columns=[
        ("part_no",         "VARCHAR(20)",  "NOT NULL","—",     "Part number (PK)"),
        ("category_id",     "INT",          "NOT NULL","—",     "Links to Part_Category"),
        ("description",     "VARCHAR(255)", "NOT NULL","—",     "Part description"),
        ("unit_of_measure", "VARCHAR(20)",  "NOT NULL","—",     "KG / PCS / MTR / SET etc."),
        ("unit_rate",       "DECIMAL(12,2)","NOT NULL","0.00",  "Standard unit cost in INR"),
        ("opening_stock",   "INT",          "NOT NULL","0",     "Opening stock at year start"),
        ("minimum_stock",   "INT",          "NOT NULL","0",     "Reorder trigger level"),
        ("order_quantity",  "INT",          "NOT NULL","0",     "Standard order lot size"),
        ("part_type",       "ENUM(RAW/FINISHED)","NOT NULL","—","Classifies raw or finished"),
    ],
    pks=["part_no"],
    fks=[("category_id","Part_Category","category_id")])

make_table_sheet(wb, "Vendor", "Vendor",
    columns=[
        ("vendor_code",   "VARCHAR(20)",  "NOT NULL","—",    "Unique vendor code (PK)"),
        ("vendor_name",   "VARCHAR(100)", "NOT NULL","—",    "Full legal name"),
        ("address_line1", "VARCHAR(255)", "NULL",   "NULL",  "Street / building address"),
        ("city",          "VARCHAR(50)",  "NULL",   "NULL",  "City"),
        ("state",         "VARCHAR(50)",  "NULL",   "NULL",  "State"),
        ("pin_code",      "VARCHAR(10)",  "NULL",   "NULL",  "Postal pin code"),
        ("phone",         "VARCHAR(15)",  "NULL",   "NULL",  "Contact phone number"),
        ("email",         "VARCHAR(100)", "NULL",   "NULL",  "Official email"),
        ("payment_terms", "VARCHAR(255)", "NULL",   "NULL",  "Net 30 / Advance etc."),
    ],
    pks=["vendor_code"], fks=[])

make_table_sheet(wb, "Vendor_Part", "Vendor_Part (Junction)",
    columns=[
        ("vendor_code",    "VARCHAR(20)",   "NOT NULL","—",   "FK → Vendor"),
        ("part_no",        "VARCHAR(20)",   "NOT NULL","—",   "FK → Part"),
        ("vendor_rate",    "DECIMAL(12,2)", "NOT NULL","—",   "Vendor-specific unit rate"),
        ("lead_time_days", "INT",           "NOT NULL","0",   "Procurement lead time"),
        ("is_preferred",   "TINYINT(1)",    "NOT NULL","0",   "1 = preferred vendor"),
    ],
    pks=["vendor_code","part_no"],
    fks=[("vendor_code","Vendor","vendor_code"),
         ("part_no","Part","part_no")])

make_table_sheet(wb, "Transporter", "Transporter",
    columns=[
        ("transporter_id",   "VARCHAR(20)",  "NOT NULL","—",   "Transporter code (PK)"),
        ("transporter_name", "VARCHAR(100)", "NOT NULL","—",   "Company name"),
        ("contact_person",   "VARCHAR(100)", "NULL",   "NULL", "Point of contact"),
        ("phone",            "VARCHAR(15)",  "NULL",   "NULL", "Phone"),
        ("email",            "VARCHAR(100)", "NULL",   "NULL", "Email"),
        ("address",          "VARCHAR(255)", "NULL",   "NULL", "Registered address"),
    ],
    pks=["transporter_id"], fks=[])

make_table_sheet(wb, "Purchase_Order", "Purchase_Order",
    columns=[
        ("po_number",              "VARCHAR(20)",  "NOT NULL","—",      "PO number (PK)"),
        ("vendor_code",            "VARCHAR(20)",  "NOT NULL","—",      "FK → Vendor"),
        ("po_date",                "DATE",         "NOT NULL","—",      "Date PO was raised"),
        ("expected_delivery_date", "DATE",         "NULL",   "NULL",    "Expected delivery date"),
        ("total_amount",           "DECIMAL(14,2)","NULL",   "0.00",   "Total PO value in INR"),
        ("status",                 "ENUM(PENDING/PARTIAL/COMPLETE/CANCELLED)","NOT NULL","PENDING","PO status"),
    ],
    pks=["po_number"],
    fks=[("vendor_code","Vendor","vendor_code")])

make_table_sheet(wb, "PO_Detail", "Purchase_Order_Detail",
    columns=[
        ("po_detail_id", "INT AUTO_INCREMENT","NOT NULL","—",  "Line item PK"),
        ("po_number",    "VARCHAR(20)",       "NOT NULL","—",  "FK → Purchase_Order"),
        ("part_no",      "VARCHAR(20)",       "NOT NULL","—",  "FK → Part"),
        ("ordered_qty",  "INT",               "NOT NULL","—",  "Quantity ordered"),
        ("unit_rate",    "DECIMAL(12,2)",     "NOT NULL","—",  "Rate at time of PO"),
    ],
    pks=["po_detail_id"],
    fks=[("po_number","Purchase_Order","po_number"),
         ("part_no","Part","part_no")])

make_table_sheet(wb, "Challan", "Challan",
    columns=[
        ("challan_no",     "VARCHAR(20)", "NOT NULL","—",   "Challan number (PK)"),
        ("challan_date",   "DATE",        "NOT NULL","—",   "Date of delivery"),
        ("vendor_code",    "VARCHAR(20)", "NOT NULL","—",   "FK → Vendor"),
        ("transporter_id", "VARCHAR(20)", "NOT NULL","—",   "FK → Transporter"),
        ("po_number",      "VARCHAR(20)", "NOT NULL","—",   "FK → Purchase_Order"),
        ("remarks",        "TEXT",        "NULL",   "NULL", "Optional remarks"),
    ],
    pks=["challan_no"],
    fks=[("vendor_code","Vendor","vendor_code"),
         ("transporter_id","Transporter","transporter_id"),
         ("po_number","Purchase_Order","po_number")])

make_table_sheet(wb, "GRR", "GRR (Goods Received Report)",
    columns=[
        ("grr_no",      "INT AUTO_INCREMENT","NOT NULL","—",   "Auto-generated GRR number (PK)"),
        ("challan_no",  "VARCHAR(20)",       "NOT NULL","—",   "FK → Challan"),
        ("grr_date",    "DATE",              "NOT NULL","—",   "Date of GRR preparation"),
        ("received_by", "VARCHAR(100)",      "NULL",   "NULL", "Name of receiving officer"),
        ("remarks",     "TEXT",              "NULL",   "NULL", "General remarks"),
    ],
    pks=["grr_no"],
    fks=[("challan_no","Challan","challan_no")])

make_table_sheet(wb, "GRR_Detail", "GRR_Detail",
    columns=[
        ("grr_detail_id", "INT AUTO_INCREMENT","NOT NULL","—",   "Line PK"),
        ("grr_no",        "INT",               "NOT NULL","—",   "FK → GRR"),
        ("part_no",       "VARCHAR(20)",       "NOT NULL","—",   "FK → Part"),
        ("challan_qty",   "INT",               "NOT NULL","—",   "Quantity as per challan"),
        ("description",   "VARCHAR(255)",      "NULL",   "NULL", "Part description (denormalised for GRR printout)"),
        ("remarks",       "TEXT",              "NULL",   "NULL", "Line-level remarks"),
    ],
    pks=["grr_detail_id"],
    fks=[("grr_no","GRR","grr_no"),
         ("part_no","Part","part_no")])

make_table_sheet(wb, "Quality_Inspection", "Quality_Inspection",
    columns=[
        ("inspection_id",   "INT AUTO_INCREMENT","NOT NULL","—",     "Inspection record PK"),
        ("grr_detail_id",   "INT",               "NOT NULL","—",     "FK → GRR_Detail"),
        ("inspection_date", "DATE",              "NOT NULL","—",     "Date of inspection"),
        ("accepted_qty",    "INT",               "NOT NULL","0",     "Quantity accepted"),
        ("rejected_qty",    "INT",               "NOT NULL","0",     "Quantity rejected"),
        ("status",          "ENUM(PASS/FAIL/PARTIAL)","NOT NULL","—","Overall inspection status"),
        ("inspector_name",  "VARCHAR(100)",      "NULL",   "NULL",   "Inspector name"),
        ("remarks",         "TEXT",              "NULL",   "NULL",   "Inspection notes"),
    ],
    pks=["inspection_id"],
    fks=[("grr_detail_id","GRR_Detail","grr_detail_id")])

make_table_sheet(wb, "MIR", "MIR (Material Issue Requisition)",
    columns=[
        ("mir_no",       "VARCHAR(20)", "NOT NULL","—",     "MIR number (PK)"),
        ("mir_date",     "DATE",        "NOT NULL","—",     "Date of requisition"),
        ("requested_by", "VARCHAR(100)","NULL",   "NULL",   "Name of requester"),
        ("department",   "VARCHAR(100)","NULL",   "NULL",   "Department requesting material"),
        ("status",       "ENUM(PENDING/ISSUED/PARTIAL/CANCELLED)","NOT NULL","PENDING","MIR status"),
    ],
    pks=["mir_no"], fks=[])

make_table_sheet(wb, "MIR_Detail", "MIR_Detail",
    columns=[
        ("mir_detail_id", "INT AUTO_INCREMENT","NOT NULL","—",  "Line PK"),
        ("mir_no",        "VARCHAR(20)",       "NOT NULL","—",  "FK → MIR"),
        ("part_no",       "VARCHAR(20)",       "NOT NULL","—",  "FK → Part"),
        ("qty_issued",    "INT",               "NOT NULL","—",  "Quantity actually issued"),
        ("issue_date",    "DATE",              "NOT NULL","—",  "Actual issue date"),
    ],
    pks=["mir_detail_id"],
    fks=[("mir_no","MIR","mir_no"),
         ("part_no","Part","part_no")])


# ═══════════════════════════════════════════════════════════════════
# SHEET – Normalization Documentation
# ═══════════════════════════════════════════════════════════════════
wsn = wb.create_sheet("Normalization")
wsn.sheet_view.showGridLines = False

wsn.merge_cells("A1:G1")
h = wsn["A1"]
h.value     = "NORMALIZATION DOCUMENTATION (1NF → 2NF → 3NF)"
h.font      = bold_font(14, WHITE)
h.fill      = hdr_fill(DARK_BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
wsn.row_dimensions[1].height = 28

nf_data = [
    ("","1NF (First Normal Form)","","","","",""),
    ("Rule","All attributes must be atomic (indivisible)","","","","",""),
    ("Rule","Each row must be uniquely identifiable","","","","",""),
    ("Violation","Un-normalised: Parts table had a multi-valued 'Suppliers' column listing multiple vendor codes in one cell.","","","","",""),
    ("Fix","Separated into Vendor_Part junction table – one row per vendor-part combination.","","","","",""),
    ("Violation","Vendor address was a single field containing city, state, pin.","","","","",""),
    ("Fix","Split into address_line1, city, state, pin_code – all atomic.","","","","",""),
    ("","","","","","",""),
    ("","2NF (Second Normal Form)","","","","",""),
    ("Rule","Must be in 1NF + No partial dependencies on a composite PK","","","","",""),
    ("Table","Vendor_Part has composite PK: (vendor_code, part_no)","","","","",""),
    ("Violation","If part description were stored in Vendor_Part, it would depend only on part_no (partial dependency).","","","","",""),
    ("Fix","Part description, unit_rate, etc. remain only in the Part table. Vendor_Part stores only vendor_rate, lead_time_days, is_preferred – all fully dependent on the composite key.","","","","",""),
    ("","","","","","",""),
    ("","3NF (Third Normal Form)","","","","",""),
    ("Rule","Must be in 2NF + No transitive dependencies (non-key → non-key)","","","","",""),
    ("Violation","Initially, Part table contained category_name and category_type.  category_name depends on category_id (non-key), not on part_no → transitive dependency.","","","","",""),
    ("Fix","Part_Category table created.  Part stores only category_id (FK). Eliminates category name/type repetition across 275 part rows.","","","","",""),
    ("Violation","GRR_Detail stored transporter_name inline – a transporter attribute that depends on transporter_id, not on grr_detail_id.","","","","",""),
    ("Fix","Transporter table created; GRR links to Challan which links to Transporter.","","","","",""),
    ("","","","","","",""),
    ("","BCNF (Boyce-Codd NF) – achieved","","","","",""),
    ("Note","Every determinant in every table is a candidate key. No anomaly remains.","","","","",""),
]

for r_idx, row_vals in enumerate(nf_data, 2):
    label = row_vals[0]
    text  = row_vals[1]
    wsn.merge_cells(f"B{r_idx}:G{r_idx}")
    c_a = wsn.cell(r_idx, 1, label)
    c_b = wsn.cell(r_idx, 2, text)

    if label == "":
        if text in ("1NF (First Normal Form)","2NF (Second Normal Form)",
                    "3NF (Third Normal Form)","BCNF (Boyce-Codd NF) – achieved"):
            c_b.font = bold_font(12, WHITE)
            wsn.merge_cells(f"A{r_idx}:G{r_idx}")
            wsn[f"A{r_idx}"].value = text
            wsn[f"A{r_idx}"].font  = bold_font(12, WHITE)
            wsn[f"A{r_idx}"].fill  = hdr_fill(MID_BLUE)
            wsn[f"A{r_idx}"].alignment = Alignment(horizontal="left",
                                                    vertical="center")
    else:
        lbl_colors = {"Rule": ("E2EFDA","375623"),
                      "Violation": (ORANGE,"833C00"),
                      "Fix": (LIGHT_BLUE,"1F3864"),
                      "Table": (YELLOW,"7F6000"),
                      "Note": ("F2F2F2","444444")}
        bg, fg = lbl_colors.get(label, (WHITE,"000000"))
        c_a.fill = hdr_fill(bg)
        c_a.font = bold_font(10, fg)
        c_a.alignment = Alignment(horizontal="center", vertical="center")
        c_a.border = thin_border()
        c_b.fill = hdr_fill(bg)
        c_b.font = norm_font(10, "000000")
        c_b.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
        c_b.border = thin_border()
    wsn.row_dimensions[r_idx].height = 28

wsn.column_dimensions["A"].width = 12
for col in "BCDEFG":
    wsn.column_dimensions[col].width = 22


# ═══════════════════════════════════════════════════════════════════
# SHEET – Primary Keys & Foreign Keys
# ═══════════════════════════════════════════════════════════════════
wsf = wb.create_sheet("PK_FK_Reference")
wsf.sheet_view.showGridLines = False

wsf.merge_cells("A1:F1")
h = wsf["A1"]
h.value     = "PRIMARY KEY & FOREIGN KEY REFERENCE"
h.font      = bold_font(14, WHITE)
h.fill      = hdr_fill(DARK_BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
wsf.row_dimensions[1].height = 28

apply_header_row(wsf, 2,
    ["Table","Column","Key Type","References Table","References Column","On Update"],
    bg=MID_BLUE)

pk_fk_data = [
    ("Part_Category",       "category_id",    "PRIMARY KEY","—","—","—"),
    ("Part",                "part_no",         "PRIMARY KEY","—","—","—"),
    ("Part",                "category_id",     "FOREIGN KEY","Part_Category","category_id","CASCADE"),
    ("Vendor",              "vendor_code",     "PRIMARY KEY","—","—","—"),
    ("Vendor_Part",         "vendor_code",     "PRIMARY KEY + FK","Vendor","vendor_code","CASCADE"),
    ("Vendor_Part",         "part_no",         "PRIMARY KEY + FK","Part","part_no","CASCADE"),
    ("Transporter",         "transporter_id",  "PRIMARY KEY","—","—","—"),
    ("Purchase_Order",      "po_number",       "PRIMARY KEY","—","—","—"),
    ("Purchase_Order",      "vendor_code",     "FOREIGN KEY","Vendor","vendor_code","CASCADE"),
    ("Purchase_Order_Detail","po_detail_id",   "PRIMARY KEY","—","—","—"),
    ("Purchase_Order_Detail","po_number",      "FOREIGN KEY","Purchase_Order","po_number","CASCADE"),
    ("Purchase_Order_Detail","part_no",        "FOREIGN KEY","Part","part_no","CASCADE"),
    ("Challan",             "challan_no",      "PRIMARY KEY","—","—","—"),
    ("Challan",             "vendor_code",     "FOREIGN KEY","Vendor","vendor_code","CASCADE"),
    ("Challan",             "transporter_id",  "FOREIGN KEY","Transporter","transporter_id","CASCADE"),
    ("Challan",             "po_number",       "FOREIGN KEY","Purchase_Order","po_number","CASCADE"),
    ("GRR",                 "grr_no",          "PRIMARY KEY","—","—","—"),
    ("GRR",                 "challan_no",      "FOREIGN KEY","Challan","challan_no","CASCADE"),
    ("GRR_Detail",          "grr_detail_id",   "PRIMARY KEY","—","—","—"),
    ("GRR_Detail",          "grr_no",          "FOREIGN KEY","GRR","grr_no","CASCADE"),
    ("GRR_Detail",          "part_no",         "FOREIGN KEY","Part","part_no","CASCADE"),
    ("Quality_Inspection",  "inspection_id",   "PRIMARY KEY","—","—","—"),
    ("Quality_Inspection",  "grr_detail_id",   "FOREIGN KEY","GRR_Detail","grr_detail_id","CASCADE"),
    ("MIR",                 "mir_no",          "PRIMARY KEY","—","—","—"),
    ("MIR_Detail",          "mir_detail_id",   "PRIMARY KEY","—","—","—"),
    ("MIR_Detail",          "mir_no",          "FOREIGN KEY","MIR","mir_no","CASCADE"),
    ("MIR_Detail",          "part_no",         "FOREIGN KEY","Part","part_no","CASCADE"),
]

for i, row_d in enumerate(pk_fk_data):
    alt = (i % 2 == 0)
    apply_data_row(wsf, i+3, row_d, alt=alt)
    if "PRIMARY" in row_d[2]:
        for c in range(1, 7):
            wsf.cell(i+3, c).fill = hdr_fill(YELLOW)
    elif "FOREIGN" in row_d[2]:
        for c in range(1, 7):
            wsf.cell(i+3, c).fill = hdr_fill(ORANGE)
    wsf.row_dimensions[i+3].height = 20

for col_i, w in enumerate([28, 28, 22, 28, 22, 14], 1):
    wsf.column_dimensions[get_column_letter(col_i)].width = w


# ═══════════════════════════════════════════════════════════════════
# SHEET – Data Dictionary
# ═══════════════════════════════════════════════════════════════════
wsdd = wb.create_sheet("Data Dictionary")
wsdd.sheet_view.showGridLines = False

wsdd.merge_cells("A1:H1")
h = wsdd["A1"]
h.value     = "DATA DICTIONARY – TATA SUPPLY CHAIN DBMS"
h.font      = bold_font(14, WHITE)
h.fill      = hdr_fill(DARK_BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
wsdd.row_dimensions[1].height = 28

apply_header_row(wsdd, 2,
    ["Table","Field","Data Type","Size","Nullable","Description","Business Rule","Sample Value"],
    bg=MID_BLUE)

dd_rows = [
    ("Part_Category","category_id","INT",11,"NO","Auto-generated unique category ID","Must be unique; auto-assigned","1"),
    ("Part_Category","category_name","VARCHAR",100,"NO","Human-readable category label","Cannot be blank","Steel Components"),
    ("Part_Category","category_type","ENUM","-","NO","RAW_MATERIAL or FINISHED_PART","Only two valid values","RAW_MATERIAL"),
    ("Part","part_no","VARCHAR",20,"NO","Alphanumeric part identifier","Format: RM-XXX or FP-XXX","RM-001"),
    ("Part","unit_rate","DECIMAL","12,2","NO","Standard cost per unit in INR","Must be > 0","85.50"),
    ("Part","opening_stock","INT",11,"NO","Stock quantity at financial year start","Must be >= 0","500"),
    ("Part","minimum_stock","INT",11,"NO","Trigger level for reorder alert","Must be < opening_stock","100"),
    ("Part","order_quantity","INT",11,"NO","Standard replenishment lot size","Must be > 0","200"),
    ("Vendor","vendor_code","VARCHAR",20,"NO","Unique vendor identifier","Format: VND-XXX","VND-001"),
    ("Vendor","payment_terms","VARCHAR",255,"YES","Net 30, Advance, etc.","Agreed payment condition","Net 30 Days"),
    ("Vendor_Part","vendor_rate","DECIMAL","12,2","NO","Vendor-specific price per unit","May differ from Part.unit_rate","82.00"),
    ("Vendor_Part","lead_time_days","INT",11,"NO","Days from PO to delivery","Must be >= 0","7"),
    ("Vendor_Part","is_preferred","TINYINT",1,"NO","1 = preferred source, 0 = alternate","Only one preferred per part recommended","1"),
    ("Transporter","transporter_id","VARCHAR",20,"NO","Unique transporter code","Format: TRP-XXX; exactly 5 records","TRP-001"),
    ("Challan","challan_no","VARCHAR",20,"NO","Vendor-issued delivery note number","Unique per delivery","CHN-001"),
    ("Challan","challan_date","DATE","-","NO","Date challan was raised","Cannot be future date","2024-01-24"),
    ("GRR","grr_no","INT",11,"NO","Auto-generated GRR number","System-generated; never manual","1"),
    ("GRR","grr_date","DATE","-","NO","Date GRR was prepared","Same day or after challan_date","2024-01-24"),
    ("GRR_Detail","challan_qty","INT",11,"NO","Units received per challan","accepted_qty + rejected_qty = challan_qty","200"),
    ("Quality_Inspection","accepted_qty","INT",11,"NO","Units accepted after QC","accepted_qty + rejected_qty must equal GRR challan_qty","198"),
    ("Quality_Inspection","rejected_qty","INT",11,"NO","Units rejected after QC","Rejected parts returned to vendor","2"),
    ("Quality_Inspection","status","ENUM","-","NO","PASS / FAIL / PARTIAL","PASS if rejected=0, FAIL if accepted=0","PARTIAL"),
    ("MIR","mir_no","VARCHAR",20,"NO","Unique MIR reference number","Format: MIR-YYYY-XXX","MIR-2024-001"),
    ("MIR","mir_date","DATE","-","NO","Date of material requisition","Cannot be future date","2024-02-01"),
    ("MIR_Detail","qty_issued","INT",11,"NO","Actual quantity issued","Must be <= available stock","50"),
]

for i, row_d in enumerate(dd_rows):
    apply_data_row(wsdd, i+3, row_d, alt=(i%2==0))
    wsdd.row_dimensions[i+3].height = 20

dd_widths = [24, 22, 14, 8, 10, 38, 44, 18]
for ci, w in enumerate(dd_widths, 1):
    wsdd.column_dimensions[get_column_letter(ci)].width = w


# ═══════════════════════════════════════════════════════════════════
# SHEET – 15 SQL Queries Reference
# ═══════════════════════════════════════════════════════════════════
wsq = wb.create_sheet("SQL Queries")
wsq.sheet_view.showGridLines = False

wsq.merge_cells("A1:E1")
h = wsq["A1"]
h.value     = "15 MEANINGFUL SQL QUERIES — PURPOSE REFERENCE"
h.font      = bold_font(14, WHITE)
h.fill      = hdr_fill(DARK_BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
wsq.row_dimensions[1].height = 28

apply_header_row(wsq, 2,
    ["#","Query Name","Tables Used","SQL Concept","Business Purpose"],
    bg=MID_BLUE)

queries = [
    (1, "Full Parts Catalogue",             "Part, Part_Category",                       "JOIN, ORDER BY",              "View all parts with category type for inventory catalogue"),
    (2, "Vendors for a Specific Part",      "Vendor, Vendor_Part",                       "JOIN, WHERE, ORDER BY",       "Find all suppliers for RM-001, sorted by cheapest rate"),
    (3, "Reorder Alert Report",             "Part",                                      "WHERE, Computed Column",      "Identify parts below minimum stock for procurement action"),
    (4, "Complete GRR Report",              "GRR, Challan, Vendor, Transporter, PO, Part","Multi-JOIN",                 "Full GRR document view for audit and record purposes"),
    (5, "QC Summary by Status",             "Quality_Inspection",                        "GROUP BY, SUM, ROUND",        "Accepted vs rejected statistics for quality dashboard"),
    (6, "Monthly Material Issued Report",   "MIR_Detail, Part, MIR",                     "JOIN, WHERE, GROUP BY, SUM",  "Total qty and value issued per part within a date range"),
    (7, "Vendor Performance Ranking",       "Vendor, Challan, GRR, GRR_Detail, QC",      "Multi-JOIN, GROUP BY, ROUND", "Rank vendors by rejection percentage to identify poor suppliers"),
    (8, "Open Purchase Orders",             "Purchase_Order, Vendor",                    "JOIN, WHERE, ORDER BY",       "All pending/partial POs sorted by overdue delivery date"),
    (9, "Cheapest Vendor per Part",         "Vendor_Part, Vendor, Part",                 "Correlated Sub-query",        "Best-price sourcing guide for procurement team"),
    (10,"Dept-wise Consumption Report",     "MIR, MIR_Detail, Part",                     "JOIN, GROUP BY, SUM",         "Track which department consumes which parts and at what cost"),
    (11,"Dead Stock Alert",                 "Part, MIR_Detail",                          "NOT IN Sub-query",            "Identify parts never issued — potential dead stock"),
    (12,"Transporter Delivery Summary",     "Transporter, Challan, GRR, GRR_Detail",     "LEFT JOIN, GROUP BY",         "Rank transporters by delivery volume and diversity"),
    (13,"Monthly Procurement Trend",        "Challan, GRR, GRR_Detail, Vendor_Part",     "JOIN, DATE_FORMAT, GROUP BY", "Plot monthly spend to detect procurement spikes"),
    (14,"Multi-Source Part Rate Variance",  "Vendor_Part, Part",                         "GROUP BY, HAVING, MIN/MAX",   "Identify cost saving opportunities via vendor rate comparison"),
    (15,"Live Stock Position Dashboard",    "Part, MIR_Detail",                          "LEFT JOIN, CASE WHEN",        "Real-time inventory with reorder status classification"),
]

for i, row_d in enumerate(queries):
    apply_data_row(wsq, i+3, row_d, alt=(i%2==0))
    wsq.row_dimensions[i+3].height = 22

for ci, w in enumerate([4,30,46,28,52], 1):
    wsq.column_dimensions[get_column_letter(ci)].width = w


# ═══════════════════════════════════════════════════════════════════
# SHEET – Assumptions
# ═══════════════════════════════════════════════════════════════════
wsa = wb.create_sheet("Assumptions")
wsa.sheet_view.showGridLines = False

wsa.merge_cells("A1:D1")
h = wsa["A1"]
h.value     = "DATABASE ASSUMPTIONS"
h.font      = bold_font(14, WHITE)
h.fill      = hdr_fill(DARK_BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
wsa.row_dimensions[1].height = 28

apply_header_row(wsa, 2, ["#","Area","Assumption","Impact"], bg=MID_BLUE)

assumptions = [
    (1, "Stock",      "Opening stock represents start of current financial year (1st April). No mid-year adjustments are modelled.","Simplifies stock tracking; a separate 'adjustments' table can be added later."),
    (2, "Vendors",    "A vendor-part rate (Vendor_Part.vendor_rate) may differ from the standard Part.unit_rate and is captured at the time of agreement.","Allows dynamic multi-vendor pricing without duplicating Part records."),
    (3, "Transporters","Exactly 5 transporters as stated in requirements. Table accommodates future expansion.","TRP-001 to TRP-005 are seeded; adding more requires no schema change."),
    (4, "GRR",        "One GRR is prepared per challan. If a vendor sends multiple challans on the same day, each gets a separate GRR.","grr_no is AUTO_INCREMENT to prevent collisions."),
    (5, "QC",         "Quality inspection is done at GRR detail (line) level, not at header level. Each part line can have its own acceptance / rejection.","Supports partial acceptance per line item."),
    (6, "MIR",        "MIR captures demand; actual issuance is confirmed by MIR_Detail.issue_date. qty_issued = actual issued (may be less than requested).","Supports partial issuance scenarios."),
    (7, "Currency",   "All monetary values are in Indian Rupees (INR).","DECIMAL(12,2) supports amounts up to 9,999,999,999.99."),
    (8, "Part Types", "Part.part_type mirrors Part_Category.category_type for faster filtering without a join in operational queries.","Controlled redundancy — both fields must be kept in sync by the application layer."),
    (9, "Deletion",   "ON DELETE CASCADE is NOT used. Parent records must be archived before child records can be deleted.","Prevents accidental loss of transactional history."),
    (10,"Audit",      "Audit trails (created_at, updated_at, modified_by) are not in scope for this project submission but should be added in production.","Add TIMESTAMP columns and trigger-based logging in production."),
]

for i, row_d in enumerate(assumptions):
    apply_data_row(wsa, i+3, row_d, alt=(i%2==0))
    wsa.row_dimensions[i+3].height = 32

for ci, w in enumerate([4,18,58,54], 1):
    wsa.column_dimensions[get_column_letter(ci)].width = w


# ═══════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════
out = "C:/Users/Pranav/Desktop/Tata_SCDB_TableStructure.xlsx"
wb.save(out)
print(f"Saved: {out}")
